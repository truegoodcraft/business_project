# TGC BUS Core (Business Utility System Core)
# Copyright (C) 2025 True Good Craft
#
# This file is part of TGC BUS Core.
#
# TGC BUS Core is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# TGC BUS Core is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with TGC BUS Core.  If not, see <https://www.gnu.org/licenses/>.

"""FastAPI router for domain CRUD backed by SQLAlchemy."""

from __future__ import annotations

import csv
import json
import time
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from .http import (
    app,
    get_db,
    _require_session,
    require_token_ctx,
    require_session_token,
)
from core.config.paths import APP_DIR, DATA_DIR, JOURNALS_DIR, IMPORTS_DIR
from core.appdb.paths import app_data_dir
from core.services.models import Attachment, Item, Task, Vendor

require_session = require_session_token

router = APIRouter(tags=["app"])


AUDIT_PATH = JOURNALS_DIR / "plugin_audit.jsonl"

for directory in (DATA_DIR, JOURNALS_DIR, IMPORTS_DIR):
    directory.mkdir(parents=True, exist_ok=True)


PREVIEW_RETENTION_SECONDS = 24 * 60 * 60


# Ensure transactions table exists for expense/revenue tracking.
def _ensure_transactions_table(db: Session) -> None:
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                type TEXT NOT NULL CHECK(type IN ('expense','revenue')),
                amount_cents INTEGER NOT NULL,
                category TEXT,
                date TEXT NOT NULL,
                notes TEXT,
                created_at TEXT NOT NULL DEFAULT (strftime('%Y-%m-%dT%H:%M:%fZ','now'))
            );
            """
        )
    )
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_transactions_date ON transactions(date);"))


# Unchanged contract for inventory dropdown: only vendors, shape [{id, name}]
@router.get("/app/vendors")
def get_vendors(
    db: Session = Depends(get_db),
    token: str = Depends(require_session),
):
    rows = db.query(Vendor).all()
    return [{"id": r.id, "name": r.name} for r in rows]


@router.post("/app/vendors")
def create_app_vendor(
    payload: VendorCreate,
    db: Session = Depends(get_db),
    token: str = Depends(require_session),
):
    vendor = Vendor(name=payload.name, contact=payload.contact)
    db.add(vendor)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="vendor_name_conflict") from exc
    db.refresh(vendor)
    return {"id": vendor.id, "name": vendor.name}


@router.post("/transactions")
def add_transaction(
    data: Dict[str, Any],
    db: Session = Depends(get_db),
    token: str = Depends(require_session),
) -> Dict[str, Any]:
    t = data.get("type")
    amount_cents = data.get("amount_cents")
    date = data.get("date")
    if t != "expense":
        raise HTTPException(status_code=400, detail="type must be 'expense' for this route")
    if not isinstance(amount_cents, int) or amount_cents >= 0:
        raise HTTPException(status_code=400, detail="amount_cents must be negative int")
    if not date:
        raise HTTPException(status_code=400, detail="date required (YYYY-MM-DD)")

    _ensure_transactions_table(db)
    db.execute(
        text(
            """
            INSERT INTO transactions (type, amount_cents, category, date, notes)
            VALUES (:type, :amount_cents, :category, :date, :notes)
            """
        ),
        data,
    )
    db.commit()
    new_id = db.execute(text("SELECT last_insert_rowid()"))
    last_row = new_id.scalar_one()
    return {"status": "saved", "id": last_row}


@router.get("/transactions")
def list_transactions(
    since: Optional[str] = None,
    limit: int = 10,
    db: Session = Depends(get_db),
    token: str = Depends(require_session),
) -> Dict[str, Any]:
    if limit < 1 or limit > 100:
        raise HTTPException(status_code=400, detail="limit must be 1..100")
    if since is not None and (len(since) != 10 or since[4] != "-" or since[7] != "-"):
        raise HTTPException(status_code=400, detail="since must be YYYY-MM-DD")

    _ensure_transactions_table(db)
    params: Dict[str, Any] = {"limit": limit}
    where_clause = ""
    if since:
        where_clause = "WHERE date >= :since"
        params["since"] = since

    sql = f"""
        SELECT id, type, amount_cents, category, date, notes, created_at
        FROM transactions
        {where_clause}
        ORDER BY date DESC, id DESC
        LIMIT :limit
    """
    rows = db.execute(text(sql), params).fetchall()
    items: List[Dict[str, Any]] = [
        {
            "id": row[0],
            "type": row[1],
            "amount_cents": row[2],
            "category": row[3],
            "date": row[4],
            "notes": row[5],
            "created_at": row[6],
        }
        for row in rows
    ]
    return {"items": items}


@router.get("/transactions/summary")
def transactions_summary(
    window: str = "30d",
    db: Session = Depends(get_db),
    token: str = Depends(require_session),
) -> Dict[str, Any]:
    if not window.endswith("d"):
        raise HTTPException(status_code=400, detail="window must be like '30d'")
    try:
        days = int(window[:-1])
        if days < 1 or days > 3650:
            raise ValueError()
    except Exception as exc:
        raise HTTPException(status_code=400, detail="invalid window") from exc

    cutoff = (date.today() - timedelta(days=days)).isoformat()
    _ensure_transactions_table(db)
    # Totals
    totals_row = db.execute(
        text(
            """
            SELECT
                COALESCE(SUM(CASE WHEN type='revenue' AND amount_cents > 0 THEN amount_cents ELSE 0 END), 0) AS in_cents,
                COALESCE(SUM(CASE WHEN type='expense' AND amount_cents < 0 THEN -amount_cents ELSE 0 END), 0) AS out_cents
            FROM transactions
            WHERE date >= :cutoff
            """
        ),
        {"cutoff": cutoff},
    ).fetchone()
    if totals_row is None:
        totals_row = (0, 0)
    in_cents = int(totals_row[0] or 0)
    out_cents = int(totals_row[1] or 0)

    # Category breakdowns (per type)
    cat_rows = db.execute(
        text(
            """
            SELECT
                type,
                COALESCE(category, 'uncategorized') AS category,
                SUM(
                  CASE
                    WHEN type='expense' THEN -amount_cents
                    ELSE amount_cents
                  END
                ) AS cents
            FROM transactions
            WHERE date >= :cutoff
            GROUP BY type, category
            ORDER BY type, category
            """
        ),
        {"cutoff": cutoff},
    ).fetchall()
    income_cats: List[Dict[str, Any]] = []
    expense_cats: List[Dict[str, Any]] = []
    for row in cat_rows:
        entry = {"name": row[1], "amount_cents": int(row[2] or 0)}
        if row[0] == "revenue":
            income_cats.append(entry)
        elif row[0] == "expense":
            expense_cats.append(entry)

    net_cents = in_cents - out_cents
    return {
        "window": window,
        "since": cutoff,
        "in_cents": in_cents,
        "out_cents": out_cents,
        "net_cents": net_cents,
        "income": {"total_cents": in_cents, "categories": income_cats},
        "expense": {"total_cents": out_cents, "categories": expense_cats},
    }


# ---- Business Profile (per-install JSON) ----
@router.get("/business_profile")
def get_business_profile() -> Dict[str, Any]:
    """Reads %LOCALAPPDATA%\\BUSCore\\business_profile.json if present; else returns minimal defaults."""
    path = app_data_dir() / "BUSCore" / "business_profile.json"  # app_data_dir already includes BUSCore; keep explicit subdir stable
    # tolerate either location (app_data_dir root or nested BUSCore) for compatibility
    candidates = [app_data_dir() / "business_profile.json", path]
    for p in candidates:
        if p.exists():
            try:
                return json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                raise HTTPException(status_code=500, detail="invalid profile file")
    return {"business_name": None, "logo_path": None}


@router.post("/business_profile")
def set_business_profile(body: Dict[str, Any]) -> Dict[str, Any]:
    """Overwrites/creates the JSON profile: business_name, logo_path, optional address fields."""
    out = {
        "business_name": body.get("business_name"),
        "logo_path": body.get("logo_path"),
        "address_line1": body.get("address_line1"),
        "address_line2": body.get("address_line2"),
        "city": body.get("city"),
        "region": body.get("region"),
        "postal_code": body.get("postal_code"),
        "country": body.get("country"),
        "phone": body.get("phone"),
        "email": body.get("email"),
    }
    path = app_data_dir() / "business_profile.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True}


def _cleanup_old_previews() -> None:
    cutoff = time.time() - PREVIEW_RETENTION_SECONDS
    for path in IMPORTS_DIR.glob("*.json"):
        try:
            if path.stat().st_mtime < cutoff:
                path.unlink()
        except FileNotFoundError:
            continue


def _normalize_cell(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    return text or None


def _dedupe_columns(pairs: List[Tuple[int, str]]) -> List[Tuple[int, str]]:
    seen: Dict[str, int] = {}
    result: List[Tuple[int, str]] = []
    for idx, name in pairs:
        base = name or f"column_{idx}"
        count = seen.get(base, 0)
        if count:
            alias = f"{base}_{count}"
        else:
            alias = base
        seen[base] = count + 1
        result.append((idx, alias))
    return result


def _parse_csv(content: bytes) -> tuple[List[str], List[Dict[str, object]]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    try:
        header = next(reader)
    except StopIteration as exc:
        raise ValueError("empty_file") from exc

    indexed: List[Tuple[int, str]] = []
    for idx, col in enumerate(header):
        if col is None:
            continue
        alias = str(col).strip()
        if alias:
            indexed.append((idx, alias))
    indexed = _dedupe_columns(indexed)
    if not indexed:
        raise ValueError("missing_columns")

    columns = indexed
    rows: List[Dict[str, object]] = []
    for raw in reader:
        record: Dict[str, object] = {}
        has_value = False
        for idx, column in columns:
            value = raw[idx] if idx < len(raw) else None
            normalized_value = _normalize_cell(value)
            if normalized_value is not None:
                has_value = True
            record[column] = normalized_value
        if has_value:
            rows.append(record)
    return [name for _, name in columns], rows


def _parse_xlsx(content: bytes) -> tuple[List[str], List[Dict[str, object]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise ValueError("missing_openpyxl") from exc

    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)

    header_pairs: List[Tuple[int, str]] = []
    for raw in rows_iter:
        if not raw:
            continue
        candidate_pairs: List[Tuple[int, str]] = []
        for idx, cell in enumerate(raw):
            normalized = _normalize_cell(cell)
            if normalized is None:
                continue
            alias = str(normalized).strip()
            if alias:
                candidate_pairs.append((idx, alias))
        if candidate_pairs:
            header_pairs = _dedupe_columns(candidate_pairs)
            break
    if not header_pairs:
        raise ValueError("missing_columns")

    columns = [name for _, name in header_pairs]

    rows: List[Dict[str, object]] = []
    for raw in rows_iter:
        if raw is None:
            continue
        record: Dict[str, object] = {}
        has_value = False
        for idx, column in header_pairs:
            value = raw[idx] if raw and idx < len(raw) else None
            normalized_value = _normalize_cell(value)
            if normalized_value is not None:
                has_value = True
            record[column] = normalized_value
        if has_value:
            rows.append(record)
    workbook.close()
    return columns, rows


def _load_table_from_upload(filename: Optional[str], content: bytes) -> tuple[List[str], List[Dict[str, object]]]:
    suffix = (Path(filename).suffix.lower() if filename else "").strip()
    if suffix == ".csv":
        return _parse_csv(content)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _parse_xlsx(content)
    raise ValueError("unsupported_format")


def _store_preview(columns: List[str], rows: List[Dict[str, object]]) -> dict:
    preview_id = uuid4().hex
    payload = {
        "id": preview_id,
        "created_at": datetime.utcnow().isoformat() + "Z",
        "columns": columns,
        "rows": rows,
    }
    path = IMPORTS_DIR / f"{preview_id}.json"
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return payload


def _load_preview(preview_id: str) -> dict:
    path = IMPORTS_DIR / f"{preview_id}.json"
    if not path.exists():
        raise FileNotFoundError
    return json.loads(path.read_text(encoding="utf-8"))


def journal_mutate(db: Session, action: str, payload: Dict[str, object]) -> None:
    entry = {
        "ts": int(time.time()),
        "action": action,
        **payload,
    }
    journal_path = JOURNALS_DIR / "bulk_import.jsonl"
    with journal_path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(entry, ensure_ascii=False) + "\n")


def _append_plugin_audit(action: str, request: Request, payload: Dict[str, object]) -> None:
    plugin_name = request.headers.get("X-Plugin-Name", "") or "unknown"
    entry = {
        "ts": int(time.time()),
        "action": action,
        "plugin": plugin_name,
        **payload,
    }
    with AUDIT_PATH.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(entry, ensure_ascii=False) + "\n")


async def _ensure_session(request: Request) -> None:
    failure = await _require_session(request)
    if failure is not None:
        raise HTTPException(status_code=failure.status_code, detail={"error": "unauthorized"})


BULK_FIELDS = {"name", "sku", "qty", "unit", "price", "vendor_id", "notes"}


def _coerce_field(field: str, value):
    if value is None:
        return None, None
    if isinstance(value, str):
        trimmed = value.strip()
        if not trimmed:
            return None, None
        value = trimmed
    if field in {"qty", "price"}:
        try:
            return float(value), None
        except (TypeError, ValueError):
            return None, f"invalid_{field}"
    if field == "vendor_id":
        try:
            number = float(value)
        except (TypeError, ValueError):
            return None, "invalid_vendor_id"
        if not number.is_integer():
            return None, "invalid_vendor_id"
        try:
            return int(number), None
        except (TypeError, ValueError):
            return None, "invalid_vendor_id"
    return value, None


def _extract_row_data(row: Dict[str, object], mapping: Dict[str, str]) -> Tuple[Dict[str, object], List[str]]:
    data: Dict[str, object] = {}
    errors: List[str] = []
    for field, column in mapping.items():
        if field not in BULK_FIELDS:
            continue
        column_value = row.get(column)
        coerced, error = _coerce_field(field, column_value)
        if error:
            errors.append(error)
            continue
        if coerced is None:
            continue
        data[field] = coerced
    return data, errors


# ---- Pydantic models -----------------------------------------------------


class VendorBase(BaseModel):
    name: str
    contact: Optional[str] = None


class BulkCommitBody(BaseModel):
    preview_id: str
    mapping: Dict[str, str]


class VendorCreate(VendorBase):
    pass


class VendorUpdate(BaseModel):
    name: Optional[str] = None
    contact: Optional[str] = None


class VendorOut(VendorBase):
    id: int
    created_at: datetime

    class Config:
        orm_mode = True


class ItemBase(BaseModel):
    vendor_id: Optional[int] = None
    sku: Optional[str] = None
    name: str
    qty: Optional[float] = 0
    unit: Optional[str] = None
    price: Optional[float] = None
    notes: Optional[str] = None


class ItemCreate(ItemBase):
    pass


class ItemUpdate(BaseModel):
    vendor_id: Optional[int] = None
    sku: Optional[str] = None
    name: Optional[str] = None
    qty: Optional[float] = None
    unit: Optional[str] = None
    price: Optional[float] = None
    notes: Optional[str] = None


class ItemOut(ItemBase):
    id: int
    created_at: datetime

    class Config:
        orm_mode = True


class TaskBase(BaseModel):
    item_id: Optional[int] = None
    title: str
    status: Optional[str] = "pending"
    due: Optional[date] = None
    notes: Optional[str] = None


class TaskCreate(TaskBase):
    pass


class TaskUpdate(BaseModel):
    item_id: Optional[int] = None
    title: Optional[str] = None
    status: Optional[str] = None
    due: Optional[date] = None
    notes: Optional[str] = None


class TaskOut(TaskBase):
    id: int
    created_at: datetime

    class Config:
        orm_mode = True


class AttachmentBase(BaseModel):
    reader_id: str
    label: Optional[str] = None


class AttachmentOut(AttachmentBase):
    id: int
    entity_type: str
    entity_id: int
    created_at: datetime

    class Config:
        orm_mode = True


# ---- Helpers -------------------------------------------------------------


def _require_vendor(db: Session, vendor_id: Optional[int]) -> None:
    if vendor_id is None:
        return
    if db.get(Vendor, vendor_id) is None:
        raise HTTPException(status_code=422, detail="vendor_not_found")


def _require_item(db: Session, item_id: Optional[int]) -> None:
    if item_id is None:
        return
    if db.get(Item, item_id) is None:
        raise HTTPException(status_code=422, detail="item_not_found")


def _require_entity(db: Session, entity_type: str, entity_id: int) -> None:
    mapping = {
        "vendor": Vendor,
        "item": Item,
        "task": Task,
    }
    model = mapping.get(entity_type)
    if model is None:
        raise HTTPException(status_code=400, detail="unsupported_entity_type")
    if db.get(model, entity_id) is None:
        raise HTTPException(status_code=404, detail="entity_not_found")


# ---- Vendor endpoints ----------------------------------------------------


@router.get("/vendors", response_model=List[VendorOut])
def list_vendors(db: Session = Depends(get_db)) -> List[Vendor]:
    return db.query(Vendor).order_by(Vendor.id.desc()).all()


@router.post("/vendors", response_model=VendorOut)
def create_vendor(payload: VendorCreate, db: Session = Depends(get_db)) -> Vendor:
    vendor = Vendor(**payload.dict())
    db.add(vendor)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="vendor_name_conflict") from exc
    db.refresh(vendor)
    return vendor


@router.put("/vendors/{vendor_id}", response_model=VendorOut)
def update_vendor(
    vendor_id: int, payload: VendorUpdate, db: Session = Depends(get_db)
) -> Vendor:
    vendor = db.get(Vendor, vendor_id)
    if vendor is None:
        raise HTTPException(status_code=404, detail="vendor_not_found")
    updates = payload.dict(exclude_unset=True)
    for field, value in updates.items():
        setattr(vendor, field, value)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="vendor_name_conflict") from exc
    db.refresh(vendor)
    return vendor


@router.delete("/vendors/{vendor_id}")
def delete_vendor(vendor_id: int, db: Session = Depends(get_db)) -> dict:
    vendor = db.get(Vendor, vendor_id)
    if vendor is None:
        raise HTTPException(status_code=404, detail="vendor_not_found")
    db.delete(vendor)
    db.commit()
    return {"ok": True}


# ---- Item endpoints ------------------------------------------------------


@router.post("/items/bulk_preview")
async def items_bulk_preview(
    request: Request,
    file: UploadFile = File(...),
    _session: None = Depends(_ensure_session),
) -> dict:
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="empty_file")

    try:
        columns, rows = _load_table_from_upload(file.filename, content)
    except ValueError as exc:
        error = str(exc)
        if error == "missing_openpyxl":
            raise HTTPException(status_code=500, detail="missing_openpyxl") from exc
        detail = error if error in {"empty_file", "missing_columns", "unsupported_format"} else "preview_failed"
        raise HTTPException(status_code=400, detail=detail) from exc

    _cleanup_old_previews()
    stored = _store_preview(columns, rows)
    _append_plugin_audit(
        "bulk_preview",
        request,
        {"preview_id": stored["id"], "filename": file.filename or ""},
    )

    preview_rows = rows[: min(len(rows), 10)]
    return {
        "preview_id": stored["id"],
        "columns": columns,
        "preview_rows": preview_rows,
        "total_rows": len(rows),
    }


@router.post("/items/bulk_commit")
async def items_bulk_commit(
    body: BulkCommitBody,
    request: Request,
    db: Session = Depends(get_db),
    _session: None = Depends(_ensure_session),
) -> dict:
    try:
        preview = _load_preview(body.preview_id)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="preview_not_found") from exc

    columns = preview.get("columns") or []
    rows = preview.get("rows") or []
    mapping = {k: v for k, v in (body.mapping or {}).items() if v}
    if "name" not in mapping:
        raise HTTPException(status_code=400, detail="name_mapping_required")

    for column in mapping.values():
        if column not in columns:
            raise HTTPException(status_code=400, detail={"error": "unknown_column", "column": column})

    results: List[Dict[str, object]] = []
    created = 0
    updated = 0
    skipped = 0

    try:
        for idx, row in enumerate(rows):
            row_data, errors = _extract_row_data(row, mapping)
            if "name" not in row_data:
                skipped += 1
                journal_mutate(
                    db,
                    "bulk_import",
                    {"row": idx, "status": "skipped", "reason": "missing_name"},
                )
                entry: Dict[str, object] = {"row": idx, "status": "skipped", "reason": "missing_name"}
                if errors:
                    entry["warnings"] = errors
                results.append(entry)
                continue

            item = None
            sku = row_data.get("sku")
            if sku:
                item = db.query(Item).filter(Item.sku == sku).one_or_none()
            if item is None:
                item = db.query(Item).filter(Item.name == row_data["name"]).one_or_none()

            status = "updated"
            if item is None:
                item = Item(
                    name=row_data["name"],
                    sku=row_data.get("sku"),
                    qty=row_data.get("qty", 0) or 0,
                    unit=row_data.get("unit"),
                    price=row_data.get("price"),
                    vendor_id=row_data.get("vendor_id"),
                    notes=row_data.get("notes"),
                )
                db.add(item)
                db.flush()
                created += 1
                status = "created"
            else:
                if "name" in row_data:
                    item.name = row_data["name"]
                if "sku" in row_data:
                    item.sku = row_data["sku"]
                if "qty" in row_data:
                    item.qty = row_data["qty"]
                if "unit" in row_data:
                    item.unit = row_data["unit"]
                if "price" in row_data:
                    item.price = row_data["price"]
                if "vendor_id" in row_data:
                    item.vendor_id = row_data["vendor_id"]
                if "notes" in row_data:
                    item.notes = row_data["notes"]
                updated += 1

            journal_payload: Dict[str, object] = {"row": idx, "status": status, "item_id": item.id}
            if errors:
                journal_payload["warnings"] = errors
            journal_mutate(db, "bulk_import", journal_payload)

            result_entry: Dict[str, object] = {"row": idx, "status": status, "item_id": item.id}
            if errors:
                result_entry["warnings"] = errors
            results.append(result_entry)

        db.commit()
    except Exception:
        db.rollback()
        raise

    try:
        (IMPORTS_DIR / f"{body.preview_id}.json").unlink()
    except FileNotFoundError:
        pass

    summary = {
        "preview_id": body.preview_id,
        "processed": len(rows),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "results": results,
    }

    _append_plugin_audit(
        "bulk_commit",
        request,
        {
            "preview_id": body.preview_id,
            "created": created,
            "updated": updated,
            "skipped": skipped,
        },
    )

    return summary


@router.get("/items", response_model=List[ItemOut])
def list_items(
    vendor_id: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
) -> List[Item]:
    query = db.query(Item)
    if vendor_id is not None:
        query = query.filter(Item.vendor_id == vendor_id)
    return query.order_by(Item.id.desc()).all()


@router.post("/items", response_model=ItemOut)
def create_item(payload: ItemCreate, db: Session = Depends(get_db)) -> Item:
    _require_vendor(db, payload.vendor_id)
    item = Item(
        vendor_id=payload.vendor_id,
        sku=payload.sku,
        name=payload.name,
        qty=payload.qty if payload.qty is not None else 0,
        unit=payload.unit,
        price=payload.price,
        notes=payload.notes,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.put("/items/{item_id}", response_model=ItemOut)
def update_item(
    item_id: int, payload: ItemUpdate, db: Session = Depends(get_db)
) -> Item:
    item = db.get(Item, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="item_not_found")
    updates = payload.dict(exclude_unset=True)
    if "vendor_id" in updates:
        _require_vendor(db, updates.get("vendor_id"))
    for field, value in updates.items():
        if field == "qty" and value is None:
            continue
        setattr(item, field, value)
    db.commit()
    db.refresh(item)
    return item


@router.delete("/items/{item_id}")
def delete_item(item_id: int, db: Session = Depends(get_db)) -> dict:
    item = db.get(Item, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="item_not_found")
    db.delete(item)
    db.commit()
    return {"ok": True}


# ---- Task endpoints ------------------------------------------------------


@router.get("/tasks", response_model=List[TaskOut])
def list_tasks(
    item_id: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
) -> List[Task]:
    query = db.query(Task)
    if item_id is not None:
        query = query.filter(Task.item_id == item_id)
    return query.order_by(Task.id.desc()).all()


@router.post("/tasks", response_model=TaskOut)
def create_task(payload: TaskCreate, db: Session = Depends(get_db)) -> Task:
    _require_item(db, payload.item_id)
    task = Task(
        item_id=payload.item_id,
        title=payload.title,
        status=payload.status or "pending",
        due=payload.due,
        notes=payload.notes,
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    return task


@router.put("/tasks/{task_id}", response_model=TaskOut)
def update_task(
    task_id: int, payload: TaskUpdate, db: Session = Depends(get_db)
) -> Task:
    task = db.get(Task, task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="task_not_found")
    updates = payload.dict(exclude_unset=True)
    if "item_id" in updates:
        _require_item(db, updates.get("item_id"))
    for field, value in updates.items():
        setattr(task, field, value)
    db.commit()
    db.refresh(task)
    return task


@router.delete("/tasks/{task_id}")
def delete_task(task_id: int, db: Session = Depends(get_db)) -> dict:
    task = db.get(Task, task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="task_not_found")
    db.delete(task)
    db.commit()
    return {"ok": True}


# ---- Attachment endpoints ------------------------------------------------


@router.get(
    "/attachments/{entity_type}/{entity_id}", response_model=List[AttachmentOut]
)
def list_attachments(
    entity_type: str, entity_id: int, db: Session = Depends(get_db)
) -> List[Attachment]:
    _require_entity(db, entity_type, entity_id)
    return (
        db.query(Attachment)
        .filter(
            Attachment.entity_type == entity_type,
            Attachment.entity_id == entity_id,
        )
        .order_by(Attachment.id.desc())
        .all()
    )


@router.post(
    "/attachments/{entity_type}/{entity_id}", response_model=AttachmentOut
)
def create_attachment(
    entity_type: str,
    entity_id: int,
    payload: AttachmentBase,
    db: Session = Depends(get_db),
) -> Attachment:
    _require_entity(db, entity_type, entity_id)
    attachment = Attachment(
        entity_type=entity_type,
        entity_id=entity_id,
        reader_id=payload.reader_id,
        label=payload.label,
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)
    return attachment


@router.delete("/attachments/{attachment_id}")
def delete_attachment(attachment_id: int, db: Session = Depends(get_db)) -> dict:
    attachment = db.get(Attachment, attachment_id)
    if attachment is None:
        raise HTTPException(status_code=404, detail="attachment_not_found")
    db.delete(attachment)
    db.commit()
    return {"ok": True}


app.include_router(
    router,
    prefix="/app",
    dependencies=[Depends(require_token_ctx)],
)


__all__ = ["router"]
